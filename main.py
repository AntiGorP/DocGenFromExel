from openpyxl import load_workbook
from docxtpl import DocxTemplate
from shutil import rmtree
from os import path, remove, mkdir
from time import sleep

sheet = load_workbook('DataList.xlsx').active

rows = sheet.max_row
cols = sheet.max_column

data = []

for i in range(1, rows + 1):
    row = {}
    for j in range(1, cols + 1):
        cell = sheet.cell(row = i, column = j)
        row.update({'val_'+str(j):cell.value})
    data.append(row)



path = path.dirname(path.abspath(__file__))
doc = path+'\\Temp.docx'

try:
    rmtree(path+'\\Results')
except:
    pass

sleep(1)

mkdir(path+'\\Results')
rpath = path+'\\Results'
for i in data:
    if i['val_2'] is not None:
        docf = DocxTemplate(doc)
        docf.render(i)
        myfile =rpath+'\\'+i['val_1']+'.docx'
        docf.save(myfile)
        
    else:
        mkdir(path+'\\Results\\'+i['val_1'])
        rpath = path+'\\Results\\'+i['val_1']
