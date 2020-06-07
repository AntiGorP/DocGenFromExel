from openpyxl import load_workbook


def DataReader_xlsx(filename):
    '''
    Функция считывает данные из .xlsx файлов
    '''
    sheet = load_workbook(filename).active

    rows = sheet.max_row
    cols = sheet.max_column

    names = []

    for i in range(1, cols + 1):
        cell = sheet.cell(row = 1, column = i)
        # Проверяем чтобы наименования столбцов не совпадали
        if cell.value is None:
            return 1, {}
        if cell.value in names:
            return 2, {}
        if ' ' in cell.value:
            return 3, {}

        names.append(cell.value)

    data = []

    for i in range(2, rows + 1):
        row = {}
        for j in range(1, cols + 1):
            cell = sheet.cell(row = i, column = j)
            row.update({names[j-1]:cell.value})
        data.append(row)
    return 0, data



def dataReader(filename):
    '''
    Функция считывает данные из файлов
    Первая строка должна содержать наименование столбца

    Функция возвращает код и список словарей для каждой строки, в качестве ключа к словарю принимается наиманование столбца

    Также создается 

    Коды функции
    0 - Все в порядке
    1 - В строке с названиями имеются пустые клетки
    2 - Имеются совпадающеие наименования в названии столбцов
    3 - Названия столбцов не должны содержать пробелы

    999 - тип файла не поддерживается
    '''

    if filename.endswith('.xlsx'):
        return DataReader_xlsx(filename)
    else: 
        return 999, {}
